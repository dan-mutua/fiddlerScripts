import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font
import logging

logging.basicConfig(filename='script.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def search_term_in_documentation(term, data):
    """Search for a term in the Microsoft documentation JSON data."""
    try:
        for item in data:
            if term.lower() in item['ips'].lower() or term.lower() in item['urls'].lower():
                return True
        return False
    except Exception as e:
        logging.error(f"Error searching term '{term}' in documentation: {e}")
        return False

def main():
    try:
        df = pd.read_excel('filtered_results.xlsx')

        # Fetch the JSON data 
        url = "https://endpoints.office.com/endpoints/worldwide?clientrequestid=b10c5ed1-bad1-445f-b386-b919946339a7"
        response = requests.get(url)
        data = response.json()

        results = []
        for index, row in df.iterrows():
            x_hostip = row['x-hostip']
            https_client_snihostname = row['https-client-snihostname']

            x_hostip_found = search_term_in_documentation(x_hostip, data)
            https_client_snihostname_found = search_term_in_documentation(https_client_snihostname, data)

            if https_client_snihostname_found and not x_hostip_found:
                status = 'Pass'
            elif not https_client_snihostname_found and x_hostip_found:
                status = 'Pass'
            elif not https_client_snihostname_found and not x_hostip_found:
                status = 'Fail'
            else:
                status = 'Pass'

            results.append({
                'Process': row['Process'],
                 'x-hostip': x_hostip,
                 'https-client-snihostname': https_client_snihostname,
                 'Status': status
        })

        # Create a DataFrame from the results
        results_df = pd.DataFrame(results)

        results_df.to_excel('search_results.xlsx', index=False)

        wb = load_workbook('search_results.xlsx')
        ws = wb.active

        pass_font = Font(bold=True, color="00FF00")
        fail_font = Font(bold=True, color="FF0000")

        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                if cell.value == 'Pass':
                    cell.font = pass_font
                elif cell.value == 'Fail':
                    cell.font = fail_font

        wb.save('search_results.xlsx')

        logging.info("The search results have been recorded in search_results.xlsx file with formatted statuses.")
        print("The search results have been recorded in search_results.xlsx file with formatted statuses.")
    except Exception as e:
        logging.error(f"Error in main function: {e}")

if __name__ == "__main__":
    main()